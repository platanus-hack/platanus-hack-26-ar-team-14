from __future__ import annotations

import base64
import csv
import io
import logging
from pathlib import Path

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

MAX_TEXT_ATTACHMENT_CHARS = 20_000
MAX_XLSX_SHEETS = 4
MAX_XLSX_ROWS = 120
MAX_XLSX_COLS = 16

IMAGE_MEDIA_TYPES = {
    "image/png",
    "image/jpeg",
    "image/jpg",
    "image/webp",
    "image/gif",
}

TEXT_EXTENSIONS = {".txt", ".md", ".markdown", ".csv"}


def _truncate_text(text: str, *, limit: int = MAX_TEXT_ATTACHMENT_CHARS) -> str:
    text = text.strip()
    if len(text) <= limit:
        return text
    return f"{text[:limit].rstrip()}\n\n[contenido truncado]"


def _decode_text_file(file_name: str, file_bytes: bytes) -> str:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        decoded = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(decoded))
        rows = [
            " | ".join(cell.strip() for cell in row if cell and cell.strip())
            for row in reader
        ]
        return _truncate_text("\n".join(row for row in rows if row))
    return _truncate_text(file_bytes.decode("utf-8-sig", errors="replace"))


def _extract_xlsx_text(file_bytes: bytes) -> str:
    workbook = load_workbook(
        filename=io.BytesIO(file_bytes),
        data_only=True,
        read_only=True,
    )
    chunks: list[str] = []
    for sheet in workbook.worksheets[:MAX_XLSX_SHEETS]:
        rows: list[str] = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index > MAX_XLSX_ROWS:
                rows.append("[filas restantes truncadas]")
                break
            values = [
                str(value).strip()
                for value in row[:MAX_XLSX_COLS]
                if value not in (None, "")
            ]
            if values:
                rows.append(" | ".join(values))
        if rows:
            chunks.append(f"Hoja: {sheet.title}\n" + "\n".join(rows))
    workbook.close()
    return _truncate_text("\n\n".join(chunks))


def _validate_supported_file(file: UploadFile) -> None:
    suffix = Path(file.filename or "").suffix.lower()
    content_type = (file.content_type or "").lower()
    if suffix == ".pdf":
        return
    if content_type in IMAGE_MEDIA_TYPES:
        return
    if suffix in TEXT_EXTENSIONS:
        return
    if suffix == ".xlsx":
        return
    raise HTTPException(
        status_code=400,
        detail=(
            f"Tipo de archivo no soportado: {file.filename}. "
            "Solo se aceptan PDF, imágenes, texto, CSV y XLSX."
        ),
    )


def _ingest_pdf_as_guia(
    db: Session,
    teacher_id: int,
    file_name: str,
    file_bytes: bytes,
) -> dict | None:
    """Auto-ingest an attached PDF into a fresh Guía for the teacher.

    Best-effort: errors are logged and swallowed so a bad PDF never takes
    down the chat turn — the document is still attached to the model
    context as a base64 blob even if ingestion fails.

    Returns a summary dict (`guia_id`, `name`, `question_count`) on
    success, or None if nothing usable was extracted.
    """
    # Local imports to keep chat_attachments importable in contexts that
    # don't have the full ORM stack ready.
    from app.models import Guia, GuiaItem
    from app.worksheets.store import ingest_pdf

    try:
        questions = ingest_pdf(db, file_name=file_name, file_bytes=file_bytes)
    except Exception:  # noqa: BLE001
        logger.exception("auto-ingest pdf falló: %s", file_name)
        return None
    if not questions:
        return None

    guia_name = Path(file_name).stem or file_name
    existing = (
        db.query(Guia)
        .filter_by(teacher_id=teacher_id, name=guia_name)
        .one_or_none()
    )
    if existing is not None:
        return {
            "guia_id": existing.id,
            "name": existing.name,
            "question_count": len(existing.items),
        }

    guia = Guia(
        teacher_id=teacher_id,
        name=guia_name,
        items=[
            GuiaItem(question_id=q.id, ordinal=i) for i, q in enumerate(questions)
        ],
    )
    db.add(guia)
    db.commit()
    db.refresh(guia)
    return {
        "guia_id": guia.id,
        "name": guia.name,
        "question_count": len(guia.items),
    }


async def build_last_user_message_content(
    text: str,
    files: list[UploadFile],
    *,
    db: Session | None = None,
    teacher_id: int | None = None,
) -> list[dict]:
    """Compose the LLM-facing content for the user's last message.

    When `db` and `teacher_id` are both provided, attached PDFs get
    auto-ingested into Question rows + a fresh Guía for the teacher.
    The Guía's id and question count are surfaced in the message text
    so the agent can attach it to the plan via `crear_material_para_plan`.
    """
    for file in files:
        _validate_supported_file(file)

    text = text.strip()
    summaries = [
        f"- {file.filename} ({Path(file.filename or '').suffix.lower() or file.content_type or 'archivo'})"
        for file in files
    ]
    intro_parts = [text] if text else []
    if summaries:
        intro_parts.append("Archivos adjuntos en este turno:\n" + "\n".join(summaries))

    # Collect ingested-guía notes here; they're appended to the intro
    # block after we walk all files, so the agent sees them right next
    # to the file list.
    guia_notes: list[str] = []

    content: list[dict] = []

    for file in files:
        file_name = file.filename or "archivo"
        suffix = Path(file_name).suffix.lower()
        content_type = (file.content_type or "").lower()
        file_bytes = await file.read()
        if not file_bytes:
            continue

        if suffix == ".pdf":
            content.append(
                {
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": base64.b64encode(file_bytes).decode("ascii"),
                    },
                }
            )
            if db is not None and teacher_id is not None:
                summary = _ingest_pdf_as_guia(
                    db,
                    teacher_id=teacher_id,
                    file_name=file_name,
                    file_bytes=file_bytes,
                )
                if summary is not None:
                    guia_notes.append(
                        f'- {file_name}: guía creada (id={summary["guia_id"]}, '
                        f'nombre="{summary["name"]}", '
                        f'{summary["question_count"]} preguntas)'
                    )
            continue

        if content_type in IMAGE_MEDIA_TYPES:
            media_type = "image/jpeg" if content_type == "image/jpg" else content_type
            content.append({"type": "text", "text": f"Imagen adjunta: {file_name}"})
            content.append(
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": media_type,
                        "data": base64.b64encode(file_bytes).decode("ascii"),
                    },
                }
            )
            continue

        if suffix in TEXT_EXTENSIONS:
            extracted = _decode_text_file(file_name, file_bytes)
        elif suffix == ".xlsx":
            extracted = _extract_xlsx_text(file_bytes)
        else:
            extracted = ""

        content.append(
            {
                "type": "text",
                "text": f"Archivo adjunto: {file_name}\n\n{extracted or '[sin texto legible]'}",
            }
        )

    if guia_notes:
        intro_parts.append(
            "Guías recién subidas (ya quedaron en el banco; pendiente "
            "asignarlas a una fila del plan si el docente lo pide):\n"
            + "\n".join(guia_notes)
        )

    intro_text = (
        "\n\n".join(part for part in intro_parts if part).strip()
        or "Revisa los archivos adjuntos y ayúdame con esta planificación."
    )
    content.insert(0, {"type": "text", "text": intro_text})

    return content
