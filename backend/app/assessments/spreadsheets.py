"""Spreadsheet conversion and normalization for assessment results."""

from __future__ import annotations

import csv
import io
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.config import settings

QUESTION_HEADER_RE = re.compile(r"(?:^|[^a-z])p(?:regunta)?\s*0*(\d+)$", re.IGNORECASE)
BARE_NUMBER_RE = re.compile(r"^0*(\d+)$")
TOTAL_HEADERS = {"total", "puntaje total", "promedio", "nota", "score", "resultado"}


@dataclass(frozen=True)
class ParsedResultsRow:
    student_name: str
    question_scores: dict[str, float]
    total_score: float | None


@dataclass(frozen=True)
class ParsedResults:
    question_keys: list[str]
    rows: list[ParsedResultsRow]


def _run_soffice(input_path: Path, target_format: str, outdir: Path) -> Path:
    if shutil.which(settings.libreoffice_bin) is None:
        raise RuntimeError(
            f"No encuentro el binario de LibreOffice: {settings.libreoffice_bin}"
        )
    subprocess.run(
        [
            settings.libreoffice_bin,
            "--headless",
            "--convert-to",
            target_format,
            "--outdir",
            str(outdir),
            str(input_path),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    candidates = sorted(outdir.glob(f"{input_path.stem}*.{target_format}"))
    if not candidates:
        raise RuntimeError(f"LibreOffice no generó salida {target_format}.")
    return candidates[0]


def convert_spreadsheet_to_pdf(file_name: str, file_bytes: bytes) -> bytes:
    suffix = Path(file_name).suffix.lower() or ".xlsx"
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        source = tmp_path / f"input{suffix}"
        source.write_bytes(file_bytes)
        output = _run_soffice(source, "pdf", tmp_path)
        return output.read_bytes()


def parse_results_file(file_name: str, file_bytes: bytes) -> ParsedResults:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        return _parse_csv(file_bytes)
    if suffix == ".xlsx":
        return _parse_xlsx(file_bytes)
    if suffix == ".xls":
        converted = _convert_xls_to_xlsx(file_bytes)
        return _parse_xlsx(converted)
    raise ValueError("Formato de resultados no soportado. Usa XLSX, XLS o CSV.")


def _convert_xls_to_xlsx(file_bytes: bytes) -> bytes:
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        source = tmp_path / "input.xls"
        source.write_bytes(file_bytes)
        output = _run_soffice(source, "xlsx", tmp_path)
        return output.read_bytes()


def _parse_csv(file_bytes: bytes) -> ParsedResults:
    decoded = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(decoded))
    rows = list(reader)
    return _parse_grid(rows)


def _parse_xlsx(file_bytes: bytes) -> ParsedResults:
    workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
    return _parse_grid(rows)


def _parse_grid(raw_rows: list[list[object | None]]) -> ParsedResults:
    rows = [[_cell_to_text(cell) for cell in row] for row in raw_rows]
    header = next((row for row in rows if any(cell for cell in row)), None)
    if header is None:
        raise ValueError("El archivo de resultados está vacío.")

    question_columns: list[tuple[int, str]] = []
    total_index: int | None = None
    for idx, cell in enumerate(header):
        key = _parse_question_header(cell)
        if key:
            question_columns.append((idx, key))
            continue
        if cell.strip().lower() in TOTAL_HEADERS:
            total_index = idx

    if not question_columns:
        raise ValueError(
            "No encontré columnas de preguntas. Usa encabezados como P1, P2, P3."
        )

    parsed_rows: list[ParsedResultsRow] = []
    header_index = rows.index(header)
    for row in rows[header_index + 1 :]:
        if not any(row):
            continue
        student_name = row[0].strip()
        if not student_name:
            continue

        scores: dict[str, float] = {}
        for idx, key in question_columns:
            value = row[idx] if idx < len(row) else ""
            if not value:
                continue
            numeric = _parse_numeric(value)
            if numeric is not None:
                scores[key] = numeric

        if not scores:
            continue

        total_score = None
        if total_index is not None and total_index < len(row):
            total_score = _parse_numeric(row[total_index])
        if total_score is None:
            total_score = sum(scores.values())

        parsed_rows.append(
            ParsedResultsRow(
                student_name=student_name,
                question_scores=scores,
                total_score=total_score,
            )
        )

    if not parsed_rows:
        raise ValueError("No encontré filas de estudiantes con puntajes legibles.")

    return ParsedResults(
        question_keys=[key for _, key in question_columns],
        rows=parsed_rows,
    )


def _cell_to_text(cell: object | None) -> str:
    if cell is None:
        return ""
    return str(cell).strip()


def _parse_question_header(value: str) -> str | None:
    candidate = value.strip()
    if not candidate:
        return None
    match = QUESTION_HEADER_RE.search(candidate)
    if match:
        return f"P{int(match.group(1))}"
    bare = BARE_NUMBER_RE.match(candidate)
    if bare:
        return f"P{int(bare.group(1))}"
    return None


def _parse_numeric(value: str) -> float | None:
    normalized = value.strip().replace(",", ".")
    if not normalized:
        return None
    try:
        return float(normalized)
    except ValueError:
        return None
